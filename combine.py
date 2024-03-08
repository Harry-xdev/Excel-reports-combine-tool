import pandas as pd
import openpyxl
import os
from pathlib import Path

import xls_to_xlsx
import func_get_filename


def remove_hidden_sheet(input_file):
	workbook = openpyxl.load_workbook(input_file)
	sheets_list = workbook.sheetnames
	for sheet_name in sheets_list:
		sheet = workbook[sheet_name]
		if sheet.sheet_state == 'hidden':
			workbook.remove(sheet)
		elif sheet_name == 'BLACK MOUNTAIN 0512' or sheet_name == 'FINAL REPORT' or sheet_name == 'ORIGINAL' :
			workbook.remove(sheet)
	workbook.save(input_file)

def production_append_all_sheet_in_excel_to_df(input_file):
	def count_sheets(input_file):
		workbook = openpyxl.load_workbook(input_file)
		sheet_names = workbook.sheetnames
		sheet_len = len(sheet_names)
		return sheet_len
	sheet_len = count_sheets(input_file)

	df_list = []
	for index in range(sheet_len):
		df = pd.read_excel(input_file, sheet_name=index, skiprows=13, usecols='A:W', nrows=11)
		df_list.append(df)
	return df_list

def combine_df_list(df_list):
	combined_df = pd.DataFrame()
	for df in df_list:
		combined_df = pd.concat([combined_df, df], ignore_index=True)
	return combined_df
	
def write_combined_df_to_excel(export_file, data):
	data.to_excel(export_file, sheet_name='Sheet', index=False)


# Get list of string of file path of input files:
def get_xlsx_file_paths(folder_path):
	folder = Path(folder_path)
	xlsx_files = list(folder.glob("*.xlsx"))
	file_path_list = [str(file) for file in xlsx_files]
	return file_path_list

# Combine all sheets in a excel file to a DataFrame and remove hidden sheets:
def combine_sheets(input):
	remove_hidden_sheet(input)
	df_list = production_append_all_sheet_in_excel_to_df(input)
	combined_df = combine_df_list(df_list)
	return combined_df

def get_df_from_input(folder_path):
	file_path_list = get_xlsx_file_paths(folder_path)
	print('List of xlsx file name: ', file_path_list)
	df_list = []
	for i in file_path_list:
		df = combine_sheets(i)
		df_list.append(df)
	return df_list
	
def combine_df_to_a_DataFrame(df_list):
	combined_df = pd.DataFrame()
	for df in df_list:
		combined_df = pd.concat([combined_df, df], ignore_index=True)
	return combined_df

def delete_all_file(folder_path):
	file_list = os.listdir(folder_path)
	for name in file_list:
		file_path = os.path.join(folder_path, name)
		if os.path.isfile(file_path): #Check if file or directory
			os.remove(file_path)

def count_xlsx_file():
	script_path = os.path.abspath(__file__)
	folder_path = os.path.dirname(script_path)
	folder = Path(folder_path)
	xlsx_files = folder.glob('*.xlsx')
	xlsx_lst = [str(i) for i in xlsx_files]
	length = len(xlsx_lst)
	return length

def remove_input_files(input_path, file_format):
	remain_xlsx = func_get_filename.get_filenames(input_path, file_format)
	for i in remain_xlsx:
		i = input_path + '/' + i
		os.remove(i)
		print(f'{i} has been removed.')


def running():
	delete_all_file('output')
	print('Clear all file.')
	folder_path = 'output'
	num = count_xlsx_file()
	export_file = 'combined_report_' + str(num) + '.xlsx'
	print('Tool starting...')
	print('Converting files format...')
	xls_to_xlsx.running()
	df_list = get_df_from_input(folder_path)
	if df_list == []:
		print('df_list is empty.')
	else:
		print('.................')
		print('.....................')
		print('..........................')
		print('...............................')
		print('Amount of file: ', len(df_list))
		combined_df = combine_df_to_a_DataFrame(df_list)
		write_combined_df_to_excel(export_file, combined_df)
		print('Combined file successful.')

running()

folder_path = 'input/test'
file_format = '*.xlsx'
remove_input_files(folder_path, file_format)